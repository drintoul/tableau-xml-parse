def to_df(fields):
    """Convert Fields list to Pandas DataFrame
    """

    import pandas as pd

    try:

        df = pd.DataFrame(fields, columns=["Field", "Hidden", "Role", "DataType", "Calculation"])
        df["DataType"] = df["DataType"].str.replace("real", "float")

        return df

    except Exception as e:

        print(e)
        return False


def to_excel(df, file, sheet):
    """Save DataFrame to File, sheet by sheet. Call repeatedly to add new sheet. Note that openpyxl must be installed
    """

    import pandas as pd

    try:

        with pd.ExcelWriter(file, engine="openpyxl", sheet_exists="replace", mode="a") as writer:
            df.to_excel(writer, sheet_name=sheet, index=None)

    except Exception:

        with pd.ExcelWriter(file, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet, index=None)


def summarize_excel(directory):
    """Combine Excel files into one file and remove duplicates
    """

    from files import list_files
    import pandas as pd

    df = pd.DataFrame()

    try:

        wbs = list_files(directory, "xlsx")

        for wb in wbs:
            new_df = pd.read_excel(wb)
            df = pd.concat([df, new_df], axis=0, ignore_index=True)

        to_excel(df, "summary.xlsx", "Summary")

        return False

    except Exception as e:

        print(e)
        return False


def colorize_and_format(filename):
    """Style Excel file
    """

    import openpyxl as xl
    from openpyxl.styles import Alignment

    # lightblue, red, orange, purple, olive, aqua, navy
    colors = ["4F81BD", "C0504D", "F79646", "8064A2", "98BB59", "4BACC6", "1F497d"]

    column_to_wrap = "E"

    wb = xl.load_workbook(filename)
    sheetnames = wb.sheetnames

    for i, ws in enumerate(sheetnames):

        wb[ws].sheet_properties.tabColor = colors[i]
        wb[ws].column_dimensions["A"].width = 80
        wb[ws].column_dimensions["B"].width = 20
        wb[ws].column_dimensions["C"].width = 20
        wb[ws].column_dimensions["D"].width = 20
        wb[ws].column_dimensions["E"].width = 120

        for cell in wb[ws][column_to_wrap]:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(filename)
        
